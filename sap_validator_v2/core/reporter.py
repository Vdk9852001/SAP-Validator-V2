"""
SAP Migration Post-Load Validator — Excel Report Generator
Accepts either a ValidationResult object or the dict format used by the dashboard.
"""

from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Union, TYPE_CHECKING

if TYPE_CHECKING:
    from core.validator import ValidationResult


def _coerce(result) -> dict:
    """
    Accept either a ValidationResult dataclass or the plain dict
    that dashboard/app.py stores in results_store.
    Always returns a plain dict with consistent keys.
    """
    if isinstance(result, dict):
        return result

    # ValidationResult dataclass → dict
    ss = result.summary_stats
    field_rows = []
    for fr in result.field_results:
        field_rows.append({
            "field":       fr.field_source,
            "field_target": fr.field_target,
            "type":        "numeric" if fr.is_numeric else "string",
            "tolerance":   fr.tolerance_used,
            "total":       fr.total_records,
            "matched":     fr.matched,
            "mismatched":  fr.mismatched,
            "miss_source": fr.missing_in_source,
            "miss_target": fr.missing_in_target,
            "match_pct":   fr.match_pct,
            "status":      fr.status,
            "mismatches":  fr.mismatch_details,
        })

    mapping = None
    if result.mapping:
        mapping = {
            "join_key":           result.mapping.join_key,
            "numeric_fields":     result.mapping.numeric_fields,
            "tolerance_map":      result.mapping.tolerance_map,
            "source_only_fields": result.mapping.source_only_fields,
            "target_only_fields": result.mapping.target_only_fields,
        }

    return {
        "name":                   Path(result.source_file).stem.upper(),
        "status":                 result.overall_status,
        "source_file":            result.source_file,
        "target_file":            result.target_file,
        "total_source_records":   result.total_source_records,
        "total_target_records":   result.total_target_records,
        "records_matched":        result.records_matched,
        "records_only_in_source": result.records_only_in_source,
        "records_only_in_target": result.records_only_in_target,
        "fields_passed":          ss["fields_passed"],
        "fields_failed":          ss["fields_failed"],
        "total_fields":           ss["total_fields_validated"],
        "pass_rate_pct":          ss["pass_rate_pct"],
        "errors":                 result.errors,
        "mapping":                mapping,
        "field_results":          field_rows,
        "run_at":                 datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def generate_excel_report(result, output_path: str) -> str:
    """
    Build a formatted Excel workbook for one validation run.
    result can be a ValidationResult dataclass or a dashboard result dict.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("pip install openpyxl")

    r = _coerce(result)

    # ── Colour palette ────────────────────────────────────────────────────────
    C_NAVY       = "FF1B3A57"
    C_WHITE      = "FFFFFFFF"
    C_GREEN      = "FF00AA44"
    C_RED        = "FFCC2200"
    C_DARK       = "FF333333"
    C_LIGHT_GRN  = "FFE6F4EA"
    C_LIGHT_RED  = "FFFCE8E6"
    C_LIGHT_GREY = "FFF5F5F5"
    C_BLUE_LIGHT = "FFE8F0FE"

    def fill(color):
        return PatternFill("solid", fgColor=color)

    def border():
        s = Side(style="thin", color="FFCCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr_cell(ws, row, col, value, bg=C_NAVY):
        c = ws.cell(row, col, value)
        c.fill = fill(bg)
        c.font = Font(bold=True, color=C_WHITE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border()
        return c

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"SAP Post-Load Validation — {r['name']}"
    c.font  = Font(bold=True, size=16, color=C_WHITE)
    c.fill  = fill(C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    # Metadata block
    overall_color = C_GREEN if r["status"] == "PASS" else C_RED
    meta = [
        ("Run Date",       r.get("run_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
        ("Source File",    r["source_file"]),
        ("Target File",    r["target_file"]),
        ("Overall Status", r["status"]),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = Font(bold=True, color="FF444444")
        cell = ws.cell(i, 2, v)
        if k == "Overall Status":
            cell.font = Font(bold=True, color=overall_color, size=12)

    # Record counts section
    ws.cell(9, 1, "Record Counts").font = Font(bold=True, size=11, color="FF222222")
    record_rows = [
        ("Source Records",   r["total_source_records"]),
        ("Target Records",   r["total_target_records"]),
        ("Keys Matched",     r["records_matched"]),
        ("Source Only",      r["records_only_in_source"]),
        ("Target Only",      r["records_only_in_target"]),
    ]
    for i, (k, v) in enumerate(record_rows, start=10):
        ws.cell(i, 1, k)
        cell = ws.cell(i, 2, v)
        if k in ("Source Only", "Target Only") and v > 0:
            cell.font = Font(bold=True, color=C_RED)

    # Stats section
    ws.cell(9, 4, "Validation Stats").font = Font(bold=True, size=11, color="FF222222")
    stat_rows = [
        ("Fields Validated", r["total_fields"]),
        ("Fields Passed",    r["fields_passed"]),
        ("Fields Failed",    r["fields_failed"]),
        ("Pass Rate",        f"{r['pass_rate_pct']}%"),
    ]
    for i, (k, v) in enumerate(stat_rows, start=10):
        ws.cell(i, 4, k)
        cell = ws.cell(i, 5, v)
        if k == "Fields Failed" and isinstance(v, int) and v > 0:
            cell.font = Font(bold=True, color=C_RED)
        if k == "Pass Rate":
            cell.font = Font(bold=True, color=C_GREEN if r["fields_failed"] == 0 else C_RED)

    # Mapping info
    if r.get("mapping"):
        m = r["mapping"]
        ws.cell(9, 7, "Auto-Detected").font = Font(bold=True, size=11, color="FF222222")
        ws.cell(10, 7, "Join Key")
        ws.cell(10, 8, m["join_key"]).font = Font(bold=True)
        ws.cell(11, 7, "Numeric Fields")
        ws.cell(11, 8, ", ".join(m.get("numeric_fields", [])) or "none")
        ws.cell(12, 7, "Source-Only Skipped")
        ws.cell(12, 8, ", ".join(m.get("source_only_fields", [])) or "none")
        ws.cell(13, 7, "Target-Only Skipped")
        ws.cell(13, 8, ", ".join(m.get("target_only_fields", [])) or "none")

    # Field results table
    ws.cell(17, 1, "Field-Level Results").font = Font(bold=True, size=11, color="FF222222")
    hdrs = ["Field", "Type", "Tolerance", "Total", "Matched",
            "Mismatched", "Miss-Source", "Miss-Target", "Match %", "Status"]
    for col, h in enumerate(hdrs, 1):
        hdr_cell(ws, 18, col, h)

    for row_i, fr in enumerate(r["field_results"], start=19):
        bg  = C_LIGHT_GRN if fr["status"] == "PASS" else C_LIGHT_RED
        tol = f"±{fr['tolerance']}" if fr.get("tolerance") is not None else "—"
        vals = [
            fr.get("field") or fr.get("field_source", ""),
            fr.get("type", ""),
            tol,
            fr["total"],
            fr["matched"],
            fr["mismatched"],
            fr["miss_source"],
            fr["miss_target"],
            f"{fr['match_pct']}%",
            fr["status"],
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row_i, col, val)
            cell.fill = fill(bg)
            cell.border = border()
            cell.alignment = Alignment(
                horizontal="center" if col > 2 else "left",
                vertical="center"
            )
            if col == len(vals):
                cell.font = Font(bold=True,
                    color=C_GREEN if fr["status"] == "PASS" else C_RED)

    col_widths = [22, 10, 10, 8, 10, 12, 12, 12, 10, 10, 24, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2+: One tab per failing field ───────────────────────────────────
    failing = [fr for fr in r["field_results"]
               if fr["status"] == "FAIL" and fr.get("mismatches")]
    for fr in failing:
        fname = (fr.get("field") or fr.get("field_source", "FIELD"))[:26]
        safe  = fname.replace("/", "_").replace("\\", "_")
        ws2   = wb.create_sheet(title=f"FAIL_{safe}")
        ws2.sheet_view.showGridLines = False

        # Tab title
        ws2.merge_cells("A1:E1")
        c2 = ws2["A1"]
        c2.value = f"Mismatches — {fname}"
        c2.font  = Font(bold=True, size=13, color=C_WHITE)
        c2.fill  = fill(C_RED)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        # Sub-header: field stats
        stats_txt = (
            f"Total: {fr['total']}  |  "
            f"Matched: {fr['matched']}  |  "
            f"Issues: {fr['mismatched'] + fr['miss_source'] + fr['miss_target']}  |  "
            f"Match: {fr['match_pct']}%"
        )
        ws2.merge_cells("A2:E2")
        c3 = ws2["A2"]
        c3.value = stats_txt
        c3.font  = Font(size=10, color="FF555555")
        c3.fill  = fill(C_LIGHT_RED)
        c3.alignment = Alignment(horizontal="center")

        # Column headers
        for col, h in enumerate(
            ["Key / Material", "Source Value", "Target Value", "Issue", "Type"], 1
        ):
            hdr_cell(ws2, 4, col, h, bg=C_DARK)

        # Mismatch rows
        for ri, rec in enumerate(fr["mismatches"], start=5):
            bg_alt = C_LIGHT_RED if ri % 2 == 0 else C_LIGHT_GREY
            row_vals = [
                rec.get("material", ""),
                rec.get("source_value", ""),
                rec.get("target_value", ""),
                rec.get("issue", ""),
                fr.get("type", ""),
            ]
            for ci, v in enumerate(row_vals, 1):
                cell = ws2.cell(ri, ci, v)
                cell.fill   = fill(bg_alt)
                cell.border = border()
                if ci == 2:
                    cell.font = Font(color=C_RED)
                if ci == 3:
                    cell.font = Font(color=C_GREEN)

        for col, w in zip("ABCDE", [24, 30, 30, 36, 10]):
            ws2.column_dimensions[col].width = w

    wb.save(output_path)
    return output_path
